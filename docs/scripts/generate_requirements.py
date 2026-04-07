#!/usr/bin/env python3
"""
要件一覧スプレッドシート生成スクリプト
案件名: イベント来場予約システム
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === 定数 ===
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Yu Gothic", size=10)
EVEN_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ODD_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADERS = ["要件ID", "分類", "機能ID", "要件名", "要件詳細", "ステータス", "優先度", "関連画面/機能", "備考"]
COL_WIDTHS = [14, 14, 10, 30, 50, 12, 10, 20, 30]

# === 要件データ ===
REQUIREMENTS = [
    # --- F-01: 予約登録 ---
    ("REQ-F-001", "機能要件", "F-01", "LINE認証ログイン",
     "お客様がわざわざ新しいアカウントを作らなくていいように、LINEアカウントでそのままログインできるようにしたい",
     "確定", "高", "ログイン画面", ""),
    ("REQ-F-002", "機能要件", "F-01", "予約情報入力",
     "予約時に氏名、生年月日、来場日、来場時間、メールアドレス、電話番号、同伴者の人数を入力してもらいたい。スマホで入力しやすいUIにしてほしい",
     "確定", "高", "予約入力画面", ""),
    ("REQ-F-003", "機能要件", "F-01", "入力値バリデーション",
     "入力ミスがあった場合はその場でエラーを表示して、どこを直せばいいかすぐ分かるようにしたい。必須項目の漏れやメールアドレスの形式チェックは最低限お願いしたい",
     "確定", "高", "予約入力画面", ""),
    ("REQ-F-004", "機能要件", "F-01", "予約確認画面表示",
     "入力した内容を確認できる画面を挟んでほしい。いきなり確定されるとお客様が不安になるので、確認してから「予約する」ボタンを押す流れにしたい",
     "確定", "高", "予約確認画面", ""),
    ("REQ-F-005", "機能要件", "F-01", "予約完了通知",
     "予約が完了したらLINEのメッセージで「予約が完了しました」と通知を飛ばしてほしい。予約内容の要約も入れてもらえると助かる",
     "確定", "高", "LINE通知", ""),
    ("REQ-F-006", "機能要件", "F-01", "QRコード発行",
     "予約が完了したらチェックイン用のQRコードを発行してほしい。当日会場でスマホを見せるだけで受付できるようにしたい",
     "確定", "高", "予約完了画面", ""),
    ("REQ-F-007", "機能要件", "F-01", "予約番号自動採番",
     "予約ごとに番号を振ってほしい。お客様から問い合わせがあったときに予約番号で特定できるようにしたい",
     "確定", "中", "予約完了画面", ""),

    # --- F-02: 予約変更 ---
    ("REQ-F-008", "機能要件", "F-02", "予約内容変更",
     "予約した後でも来場日や来場時間、同伴者の人数を変更できるようにしたい。急な予定変更はよくあることなので",
     "確定", "高", "予約変更画面", ""),
    ("REQ-F-009", "機能要件", "F-02", "変更期限制御",
     "さすがに当日の変更は運営側が対応しきれないので、前日までしか変更できないようにしたい。期限を過ぎたら変更ボタンを押せなくするイメージ",
     "検討中", "中", "予約変更画面", "期限の詳細は要検討"),
    ("REQ-F-010", "機能要件", "F-02", "変更完了通知",
     "変更が完了したらLINEで変更後の内容を通知してほしい。ちゃんと変更できたか不安になるお客様もいるので",
     "確定", "高", "LINE通知", ""),

    # --- 非機能要件 ---
    ("REQ-NF-001", "非機能要件", "共通", "対応ブラウザ",
     "お客様はLINEアプリの中で操作するので、iPhoneとAndroidのLINE内ブラウザで問題なく動くようにしてほしい",
     "確定", "高", "全画面", ""),
    ("REQ-NF-002", "非機能要件", "共通", "レスポンス時間",
     "画面がなかなか表示されないとお客様が離脱してしまうので、3秒以内にはページが開くようにしてほしい",
     "確定", "中", "全画面", ""),
    ("REQ-NF-003", "非機能要件", "共通", "同時接続数",
     "イベント告知直後にアクセスが集中することがあるので、最低100人くらいは同時にアクセスしても大丈夫なようにしたい",
     "検討中", "中", "システム全体", "負荷要件は要検討"),
    ("REQ-NF-004", "非機能要件", "共通", "可用性",
     "予約受付期間中にシステムが落ちるとクレームになるので、できるだけ安定して稼働してほしい。99.5%以上の稼働率が理想",
     "検討中", "中", "システム全体", ""),
    ("REQ-NF-005", "非機能要件", "共通", "データバックアップ",
     "お客様の予約データが消えたら取り返しがつかないので、毎日バックアップを取ってほしい",
     "確定", "高", "システム全体", ""),
    ("REQ-NF-006", "非機能要件", "共通", "個人情報保護",
     "お客様の氏名や生年月日などの個人情報を扱うので、データベースに保存するときは暗号化してほしい。情報漏えいは絶対に避けたい",
     "確定", "高", "データベース", ""),
    ("REQ-NF-007", "非機能要件", "共通", "SSL/TLS通信",
     "通信は全部HTTPSで暗号化してほしい。お客様が安心して使えるように",
     "確定", "高", "システム全体", ""),
    ("REQ-NF-008", "非機能要件", "共通", "アクセスログ",
     "何か問題が起きたときに原因を追えるように、操作のログを残しておいてほしい。90日くらい保持してもらえれば十分",
     "確定", "中", "システム全体", ""),
    ("REQ-NF-009", "非機能要件", "共通", "レスポンシブデザイン",
     "基本的にスマホで使うので、スマホの画面サイズに最適化したデザインにしてほしい。PCの管理画面はPC向けでOK",
     "確定", "高", "全画面", ""),
    ("REQ-NF-010", "非機能要件", "共通", "文字コード",
     "文字化けが起きないようにUTF-8で統一してほしい。日本語の氏名で文字化けすると困るので",
     "確定", "高", "システム全体", ""),

]


def create_requirements_sheet(wb):
    """要件一覧シートを作成"""
    ws = wb.active
    ws.title = "要件一覧"

    # ヘッダー行
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # データ行
    for row_idx, req in enumerate(REQUIREMENTS, 2):
        fill = EVEN_ROW_FILL if row_idx % 2 == 0 else ODD_ROW_FILL
        for col_idx, value in enumerate(req, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col_idx in (1, 2, 3, 6, 7):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT_WRAP

    # フリーズペイン（ヘッダー行固定）
    ws.freeze_panes = "A2"

    # オートフィルター
    last_col = get_column_letter(len(HEADERS))
    last_row = len(REQUIREMENTS) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # データバリデーション（ドロップダウン）
    dv_status = DataValidation(type="list", formula1='"確定,未確定,検討中"', allow_blank=False)
    dv_status.prompt = "ステータスを選択してください"
    dv_status.promptTitle = "ステータス"
    ws.add_data_validation(dv_status)
    dv_status.add(f"F2:F{last_row}")

    dv_priority = DataValidation(type="list", formula1='"高,中,低"', allow_blank=False)
    dv_priority.prompt = "優先度を選択してください"
    dv_priority.promptTitle = "優先度"
    ws.add_data_validation(dv_priority)
    dv_priority.add(f"G2:G{last_row}")

    dv_category = DataValidation(type="list", formula1='"機能要件,非機能要件"', allow_blank=False)
    dv_category.prompt = "分類を選択してください"
    dv_category.promptTitle = "分類"
    ws.add_data_validation(dv_category)
    dv_category.add(f"B2:B{last_row}")

    # 行の高さ設定
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = 30


def create_metadata_sheet(wb):
    """文書情報シートを作成"""
    ws = wb.create_sheet(title="文書情報")
    today = date.today().strftime("%Y/%m/%d")

    metadata = [
        ("文書名", "要件一覧"),
        ("案件名", "イベント来場予約システム"),
        ("バージョン", "1.0"),
        ("作成者", ""),
        ("作成日", today),
        ("最終更新日", today),
        ("承認者", ""),
        ("承認日", ""),
        ("ステータス", "ドラフト"),
    ]

    # ヘッダースタイル
    for row_idx, (label, value) in enumerate(metadata, 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(name="Yu Gothic", bold=True, size=10)
        label_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        label_cell.border = THIN_BORDER
        label_cell.alignment = CENTER

        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = DATA_FONT
        value_cell.border = THIN_BORDER
        value_cell.alignment = LEFT_WRAP

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40


def main():
    wb = Workbook()
    create_requirements_sheet(wb)
    create_metadata_sheet(wb)

    output_dir = os.path.join(os.path.dirname(__file__), "..", "requirements")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "要件一覧_イベント来場予約システム.xlsx")
    wb.save(output_path)
    print(f"生成完了: {output_path}")
    print(f"要件数: {len(REQUIREMENTS)}件")


if __name__ == "__main__":
    main()
