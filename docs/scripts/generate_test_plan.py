#!/usr/bin/env python3
"""
テスト計画書生成スクリプト
案件名: イベント来場予約システム
サンプル画像の形式に準拠
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === スタイル定数 ===
TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
TITLE_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=14)
META_LABEL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
META_LABEL_FONT = Font(name="Yu Gothic", bold=True, size=9, color="333333")
META_VALUE_FONT = Font(name="Yu Gothic", size=9)
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=9)
DATA_FONT = Font(name="Yu Gothic", size=9)
EVEN_ROW_FILL = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
ODD_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# === 列定義 ===
HEADERS = ["No.", "要件区分", "要件名", "ステータス", "担当者", "期間(日)", "テスト種別", "テスト優先度", "テストレベル", "テスト観点"]
COL_WIDTHS = [6, 12, 35, 12, 14, 10, 16, 12, 18, 50]

# === テスト計画データ（要件定義書の20件に基づく） ===
TEST_PLAN_DATA = [
    # --- 機能要件 10件 ---
    (1, "機能要件", "LINE認証ログイン",
     "未着手", "mihayang_cho", 2, "結合テスト",
     "高", "結合テスト→システムテスト",
     "①LINE未ログイン状態からの認証フロー\n②認証トークン期限切れ時の再認証\n③LINE連携解除済みアカウントでのアクセス"),
    (2, "機能要件", "予約情報入力",
     "未着手", "mihayang_cho", 3, "結合テスト",
     "高", "単体テスト→結合テスト",
     "①全項目を正しく入力した場合の登録成功\n②必須項目を未入力にした場合のエラー表示\n③スマホ画面での入力操作性（キーボード切替・日付選択）"),
    (3, "機能要件", "入力値バリデーション",
     "未着手", "mihayang_cho", 2, "単体テスト",
     "高", "単体テスト→結合テスト",
     "①メールアドレス形式不正時のエラーメッセージ表示\n②生年月日の未来日入力の拒否\n③電話番号の桁数・形式チェック"),
    (4, "機能要件", "予約確認画面表示",
     "未着手", "", 1, "結合テスト",
     "中", "結合テスト",
     "①入力画面の内容が確認画面に正しく反映されるか\n②「戻る」操作で入力内容が保持されるか"),
    (5, "機能要件", "予約完了通知",
     "未着手", "", 2, "結合テスト",
     "高", "結合テスト→システムテスト",
     "①予約確定直後にLINE通知が届くか\n②通知メッセージに予約内容（日時・予約番号）が含まれるか\n③通信エラー時の通知リトライ"),
    (6, "機能要件", "QRコード発行",
     "未着手", "", 2, "結合テスト",
     "高", "結合テスト→システムテスト",
     "①発行されたQRコードがスマホ画面上で読み取り可能か\n②予約ごとに異なるQRコードが生成されるか"),
    (7, "機能要件", "予約番号自動採番",
     "未着手", "", 1, "単体テスト",
     "中", "単体テスト",
     "①同時に複数予約が入った場合に番号が重複しないか\n②採番された番号で予約を一意に特定できるか"),
    (8, "機能要件", "予約内容変更",
     "未着手", "", 2, "結合テスト",
     "高", "結合テスト→システムテスト",
     "①来場日・時間・同伴者人数それぞれの変更が反映されるか\n②変更後に定員超過にならないか\n③他人の予約を変更できないこと"),
    (9, "機能要件", "変更期限制御",
     "未着手", "", 1, "結合テスト",
     "中", "結合テスト",
     "①来場日前日23:59まで変更ボタンが有効か\n②期限超過後に変更操作がブロックされるか"),
    (10, "機能要件", "変更完了通知",
     "未着手", "", 1, "結合テスト",
     "高", "結合テスト→システムテスト",
     "①変更確定後にLINE通知が届くか\n②通知に変更後の内容が正しく記載されているか"),

    # --- 非機能要件 10件 ---
    (11, "非機能要件", "対応ブラウザ",
     "未着手", "", 3, "システムテスト",
     "高", "システムテスト",
     "①iOSのLINE内ブラウザで全画面が正常表示されるか\n②AndroidのLINE内ブラウザで全画面が正常表示されるか\n③OS・LINEバージョン違いによる表示差異"),
    (12, "非機能要件", "レスポンス時間",
     "未着手", "", 2, "性能テスト",
     "中", "システムテスト",
     "①主要画面遷移（予約入力→確認→完了）が各3秒以内か\n②通常負荷時と高負荷時のレスポンス差"),
    (13, "非機能要件", "同時接続数",
     "未着手", "", 2, "負荷テスト",
     "中", "システムテスト",
     "①100名同時アクセスでエラーが発生しないか\n②同時予約操作時のデータ不整合が起きないか"),
    (14, "非機能要件", "可用性",
     "未着手", "", 1, "システムテスト",
     "中", "システムテスト",
     "①長時間連続稼働でメモリリーク等が発生しないか"),
    (15, "非機能要件", "データバックアップ",
     "未着手", "", 1, "運用テスト",
     "高", "システムテスト",
     "①日次バックアップが自動実行されるか\n②バックアップからのデータ復元が正常に行えるか"),
    (16, "非機能要件", "個人情報保護",
     "未着手", "", 2, "セキュリティテスト",
     "高", "システムテスト",
     "①DB上で氏名・生年月日等が暗号化されて保存されているか\n②APIレスポンスに不要な個人情報が含まれていないか"),
    (17, "非機能要件", "SSL/TLS通信",
     "未着手", "", 1, "セキュリティテスト",
     "高", "システムテスト",
     "①全エンドポイントがHTTPSのみでアクセス可能か\n②TLS1.2未満のプロトコルが無効化されているか"),
    (18, "非機能要件", "アクセスログ",
     "未着手", "", 1, "運用テスト",
     "中", "システムテスト",
     "①ユーザー操作（予約・変更・キャンセル）がログに記録されるか\n②90日経過後のログが自動削除されるか"),
    (19, "非機能要件", "レスポンシブデザイン",
     "未着手", "", 2, "システムテスト",
     "高", "システムテスト",
     "①小型スマホ（375px幅）でレイアウト崩れがないか\n②大型スマホ（428px幅）での表示\n③フォントサイズ拡大時の表示崩れ"),
    (20, "非機能要件", "文字コード",
     "未着手", "", 1, "システムテスト",
     "高", "単体テスト→システムテスト",
     "①日本語氏名（漢字・ひらがな・カタカナ）の登録・表示で文字化けしないか\n②絵文字や環境依存文字の入力時の挙動"),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "テスト計画書"

    # === タイトル行（行1） ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1, value="テスト計画書")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 35

    # === メタ情報行（行2） ===
    today = date.today().strftime("%Y/%m/%d")
    meta_items = [
        ("作成日:", today),
        ("作成者:", "mihayang_cho"),
        ("対象案件名:", "イベント来場予約システム"),
        ("メモ:", "コメント/フィードバックがあれば記入"),
    ]

    col = 1
    for label, value in meta_items:
        lc = ws.cell(row=2, column=col, value=label)
        lc.font = META_LABEL_FONT
        lc.fill = META_LABEL_FILL
        lc.alignment = CENTER
        lc.border = THIN_BORDER

        vc = ws.cell(row=2, column=col + 1, value=value)
        vc.font = META_VALUE_FONT
        vc.alignment = LEFT_WRAP
        vc.border = THIN_BORDER

        # メモ欄は残りの列を結合
        if label == "メモ:":
            ws.merge_cells(start_row=2, start_column=col + 1, end_row=2, end_column=len(HEADERS))
        col += 2

    ws.row_dimensions[2].height = 22

    # === ヘッダー行（行3） ===
    HEADER_ROW = 3
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 25

    # === データ行（行4〜） ===
    DATA_START = 4
    for row_offset, row_data in enumerate(TEST_PLAN_DATA):
        row_idx = DATA_START + row_offset
        fill = EVEN_ROW_FILL if row_offset % 2 == 0 else ODD_ROW_FILL

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col_idx in (1, 2, 4, 6, 7, 8):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT_WRAP

        ws.row_dimensions[row_idx].height = 30

    last_row = DATA_START + len(TEST_PLAN_DATA) - 1

    # === フリーズペイン（ヘッダー固定） ===
    ws.freeze_panes = f"A{DATA_START}"

    # === オートフィルター ===
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(HEADERS))}{last_row}"

    # === データバリデーション ===
    dv_status = DataValidation(type="list", formula1='"未着手,対応中,対応済,保留"', allow_blank=False)
    dv_status.prompt = "ステータスを選択"
    ws.add_data_validation(dv_status)
    dv_status.add(f"D{DATA_START}:D{last_row}")

    dv_priority = DataValidation(type="list", formula1='"高,中,低"', allow_blank=False)
    dv_priority.prompt = "優先度を選択"
    ws.add_data_validation(dv_priority)
    dv_priority.add(f"H{DATA_START}:H{last_row}")

    dv_category = DataValidation(type="list", formula1='"機能要件,非機能要件"', allow_blank=False)
    ws.add_data_validation(dv_category)
    dv_category.add(f"B{DATA_START}:B{last_row}")

    # === 出力 ===
    output_dir = os.path.join(os.path.dirname(__file__), "..", "test_plan")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "テスト計画書_イベント来場予約システム.xlsx")
    wb.save(output_path)
    print(f"生成完了: {output_path}")
    print(f"テスト計画件数: {len(TEST_PLAN_DATA)}件")


if __name__ == "__main__":
    main()
